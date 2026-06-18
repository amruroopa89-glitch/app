import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_report(summary, steps, output_path):
    """
    Generates a beautifully formatted Excel report for the Appium E2E test run.
    """
    # Ensure reports directory exists
    dir_name = os.path.dirname(output_path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)

    wb = Workbook()
    
    # -------------------------------------------------------------
    # Sheet 1: DASHBOARD SUMMARY
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Dashboard Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_family = "Arial"
    
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1B5E20")
    bold_font = Font(name=font_family, size=10, bold=True)
    normal_font = Font(name=font_family, size=10)
    note_font = Font(name=font_family, size=9, italic=True, color="555555")

    green_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    light_green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    light_grey_fill = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    green_border_side = Side(style="thin", color="A5D6A7")
    green_border = Border(left=green_border_side, right=green_border_side, top=green_border_side, bottom=green_border_side)

    grey_border_side = Side(style="thin", color="CFD8DC")
    grey_border = Border(left=grey_border_side, right=grey_border_side, top=grey_border_side, bottom=grey_border_side)

    thin_border_side = Side(style="thin", color="E0E0E0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Title Banner
    ws_summary.merge_cells("A2:F2")
    ws_summary.row_dimensions[2].height = 40
    title_cell = ws_summary["A2"]
    title_cell.value = "GREEN HARVEST BUDDY - E2E MOBILE TEST REPORT"
    title_cell.font = title_font
    title_cell.fill = green_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Section 1: Execution Metadata
    ws_summary["A4"].value = "Execution Metadata"
    ws_summary["A4"].font = section_font

    start_dt = datetime.fromtimestamp(summary.get("startTime", 0) / 1000.0).strftime('%Y-%m-%d %H:%M:%S')
    duration_sec = (summary.get("endTime", 0) - summary.get("startTime", 0)) / 1000.0

    meta_rows = [
        ("Execution Date", start_dt),
        ("Platform Name", summary.get("platformName", "Android")),
        ("Device Name", summary.get("deviceName", "Android Device")),
        ("Tested Browser", summary.get("browserName", "Chrome")),
        ("Target URL", summary.get("targetUrl", "")),
        ("Duration", f"{duration_sec:.2f} seconds")
    ]

    for idx, (label, val) in enumerate(meta_rows):
        r_num = 5 + idx
        c_label = ws_summary.cell(row=r_num, column=1, value=label)
        c_val = ws_summary.cell(row=r_num, column=2, value=val)
        
        c_label.font = bold_font
        c_label.fill = light_green_fill
        c_label.border = green_border
        
        c_val.font = normal_font
        c_val.border = green_border

    # Section 2: Test Metrics
    ws_summary["D4"].value = "Test Suite Metrics"
    ws_summary["D4"].font = section_font

    total_steps = len(steps)
    passed_steps = len([s for s in steps if s.get("status") == "PASS"])
    failed_steps = total_steps - passed_steps
    pass_rate = (passed_steps / total_steps * 100.0) if total_steps > 0 else 0.0

    metrics_rows = [
        ("Total Test Steps", total_steps, "3F51B5"),
        ("Passed Steps", passed_steps, "2E7D32"),
        ("Failed Steps", failed_steps, "C62828"),
        ("Pass Rate", f"{pass_rate:.1f}%", "FF9800" if pass_rate < 100 else "2E7D32")
    ]

    for idx, (label, val, color) in enumerate(metrics_rows):
        r_num = 5 + idx
        c_label = ws_summary.cell(row=r_num, column=4, value=label)
        c_val = ws_summary.cell(row=r_num, column=5, value=val)
        
        c_label.font = bold_font
        c_label.fill = light_grey_fill
        c_label.border = grey_border
        
        c_val.font = Font(name=font_family, size=11, bold=True, color=color)
        c_val.alignment = Alignment(horizontal="center")
        c_val.border = grey_border

    # Summary Description Note
    ws_summary.merge_cells("A13:F14")
    note_cell = ws_summary["A13"]
    note_cell.value = "Note: This report provides end-to-end regression validation metrics. Detailed execution steps, screenshots for each action, and failure trace statements are available in the 'Detailed Log' worksheet tab below."
    note_cell.font = note_font
    note_cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    ws_summary.column_dimensions["A"].width = 24
    ws_summary.column_dimensions["B"].width = 32
    ws_summary.column_dimensions["C"].width = 4
    ws_summary.column_dimensions["D"].width = 24
    ws_summary.column_dimensions["E"].width = 16
    ws_summary.column_dimensions["F"].width = 16

    # -------------------------------------------------------------
    # Sheet 2: DETAILED EXECUTION LOG
    # -------------------------------------------------------------
    ws_details = wb.create_sheet("Detailed Log")
    ws_details.views.sheetView[0].showGridLines = True

    headers = [
        ("ID", 10),
        ("Test Module / Feature", 22),
        ("Step Description", 35),
        ("Action Taken", 35),
        ("Expected Outcome", 35),
        ("Actual Result / Output", 45),
        ("Status", 12),
        ("Duration (ms)", 14),
        ("Timestamp", 22),
        ("Screenshot Link", 20)
    ]

    ws_details.row_dimensions[1].height = 30

    # Write headers
    for c_idx, (text, width) in enumerate(headers, 1):
        col_letter = get_column_letter(c_idx)
        ws_details.column_dimensions[col_letter].width = width
        cell = ws_details.cell(row=1, column=c_idx, value=text)
        cell.font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=Side(style="medium", color="1B5E20"), bottom=Side(style="medium", color="1B5E20"))

    # Write step logs
    for r_idx, step in enumerate(steps, 2):
        ws_details.row_dimensions[r_idx].height = 24
        
        ws_details.cell(row=r_idx, column=1, value=step.get("id", "")).alignment = Alignment(horizontal="center", vertical="center")
        ws_details.cell(row=r_idx, column=2, value=step.get("module", "")).alignment = Alignment(vertical="center")
        ws_details.cell(row=r_idx, column=3, value=step.get("description", "")).alignment = Alignment(vertical="center", wrap_text=True)
        ws_details.cell(row=r_idx, column=4, value=step.get("action", "")).alignment = Alignment(vertical="center", wrap_text=True)
        ws_details.cell(row=r_idx, column=5, value=step.get("expected", "")).alignment = Alignment(vertical="center", wrap_text=True)
        ws_details.cell(row=r_idx, column=6, value=step.get("actual", "")).alignment = Alignment(vertical="center", wrap_text=True)
        
        # Status column
        status_val = step.get("status", "")
        status_cell = ws_details.cell(row=r_idx, column=7, value=status_val)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        if status_val == "PASS":
            status_cell.fill = PatternFill(start_color="E8F5E9", fill_type="solid")
            status_cell.font = Font(name=font_family, size=10, bold=True, color="2E7D32")
        else:
            status_cell.fill = PatternFill(start_color="FFEBEE", fill_type="solid")
            status_cell.font = Font(name=font_family, size=10, bold=True, color="C62828")

        # Duration & Timestamp
        dur_cell = ws_details.cell(row=r_idx, column=8, value=step.get("duration", 0))
        dur_cell.alignment = Alignment(horizontal="right", vertical="center")
        dur_cell.number_format = '#,##0'
        
        # Convert timestamp to HH:MM:SS
        ts_str = step.get("timestamp", "")
        if ts_str:
            try:
                # ISO format parse
                dt = datetime.strptime(ts_str.split(".")[0].replace("Z", ""), "%Y-%m-%dT%H:%M:%S")
                ts_val = dt.strftime("%H:%M:%S")
            except:
                ts_val = ts_str
        else:
            ts_val = ""
        
        ws_details.cell(row=r_idx, column=9, value=ts_val).alignment = Alignment(horizontal="center", vertical="center")

        # Screenshot link
        ss_path = step.get("screenshot")
        ss_cell = ws_details.cell(row=r_idx, column=10)
        ss_cell.alignment = Alignment(horizontal="center", vertical="center")
        if ss_path:
            # We want the hyperlink relative to the report directory
            rel_name = os.path.basename(ss_path)
            link_formula = f'=HYPERLINK("./screenshots/{rel_name}", "View Image")'
            ss_cell.value = link_formula
            ss_cell.font = Font(name=font_family, size=10, color="1565C0", underline="single")
        else:
            ss_cell.value = "N/A"
            ss_cell.font = Font(name=font_family, size=10, color="9E9E9E")

        # Apply basic thin border to each cell in the detail row
        for col in range(1, 11):
            ws_details.cell(row=r_idx, column=col).border = thin_border

    # Save to file path
    wb.save(output_path)
    print(f"[+] Excel report compiled successfully at: {output_path}")
