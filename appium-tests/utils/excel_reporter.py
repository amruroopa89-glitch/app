"""
Green Harvest Buddy — Appium Excel Report Generator
Generates a professional multi-tab .xlsx with:
  Tab 1: Dashboard Summary  (KPIs, metadata, deployable status)
  Tab 2: UI/UX Tests
  Tab 3: Functional Tests
  Tab 4: Unit/Component Tests
  Tab 5: Validation Tests
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Colour constants ─────────────────────────────────────────────────────────
_C = {
    "dk_green":   "1B5E20",
    "green":      "2E7D32",
    "lt_green":   "E8F5E9",
    "mint":       "A5D6A7",
    "grey":       "ECEFF1",
    "grey_bdr":   "CFD8DC",
    "row_bdr":    "E0E0E0",
    "white":      "FFFFFF",
    "red":        "C62828",
    "lt_red":     "FFEBEE",
    "amber":      "FF8F00",
    "blue":       "1565C0",
    "text":       "212121",
    "sub":        "555555",
}

_FONT = "Calibri"


def _fill(colour):
    return PatternFill(start_color=colour, end_color=colour, fill_type="solid")


def _border(colour):
    side = Side(style="thin", color=colour)
    return Border(left=side, right=side, top=side, bottom=side)


def _med_border():
    s = Side(style="medium", color=_C["dk_green"])
    return Border(top=s, bottom=s)


def _font(**kw):
    return Font(name=_FONT, size=kw.get("size", 10), bold=kw.get("bold", False),
                italic=kw.get("italic", False),
                color=kw.get("color", _C["text"]))


# ═══════════════════════════════════════════════════════════════════════════════
def generate_excel_report(summary: dict, steps: list, output_path: str):
    """Create the multi-tab Excel report."""

    dir_name = os.path.dirname(output_path)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)

    wb = Workbook()

    # ── Categorise ────────────────────────────────────────────────────────────
    groups = {
        "ui":   [s for s in steps if "-UI" in s["id"]],
        "func": [s for s in steps if "-FUNC" in s["id"]],
        "unit": [s for s in steps if "-UNIT" in s["id"]],
        "val":  [s for s in steps if "-VAL" in s["id"]],
    }
    cats = [
        ("UI / UX Testing (Mobile)",         "ui",   groups["ui"]),
        ("Functional Testing (Mobile)",      "func", groups["func"]),
        ("Unit / Component Testing (Mobile)","unit", groups["unit"]),
        ("Validation Testing (Mobile)",      "val",  groups["val"]),
    ]

    total = len(steps)
    passed = sum(1 for s in steps if s.get("status") == "PASS")
    failed = total - passed
    rate = f"{(passed / total * 100):.1f}" if total else "0.0"
    dur_s = (summary.get("endTime", 0) - summary.get("startTime", 0)) / 1000.0
    start_dt = datetime.fromtimestamp(summary.get("startTime", 0) / 1000.0).strftime("%Y-%m-%d %H:%M:%S")

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — DASHBOARD SUMMARY
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Dashboard Summary"

    # Banner
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 42
    b = ws["A2"]
    b.value = "📱  GREEN HARVEST BUDDY — APPIUM MOBILE E2E TEST REPORT"
    b.font = _font(size=17, bold=True, color=_C["white"])
    b.fill = _fill(_C["green"])
    b.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:K3")
    ws.row_dimensions[3].height = 18
    s3 = ws["A3"]
    s3.value = f"400 Test Cases  ·  UI/UX · Functional · Unit · Validation  ·  Generated {start_dt}"
    s3.font = _font(size=9, italic=True, color=_C["white"])
    s3.fill = _fill(_C["dk_green"])
    s3.alignment = Alignment(horizontal="center", vertical="center")

    def sec(cell, label):
        ws[cell].value = label
        ws[cell].font = _font(size=11, bold=True, color=_C["dk_green"])

    # ── Execution Metadata ────────────────────────────────────────────────────
    sec("A5", "📋  Execution Metadata")
    meta = [
        ("Execution Date",  start_dt),
        ("Platform",        summary.get("platformName", "Android")),
        ("Device",          summary.get("deviceName", "Android Emulator")),
        ("Browser",         summary.get("browserName", "Chrome")),
        ("Target URL",      summary.get("targetUrl", "")),
        ("Mode",            "Appium UiAutomator2 (CI/CD)"),
        ("Duration",        f"{dur_s:.2f} seconds"),
    ]
    for i, (label, val) in enumerate(meta):
        r = 6 + i
        ws.row_dimensions[r].height = 18
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _font(bold=True); lc.fill = _fill(_C["lt_green"]); lc.border = _border(_C["mint"])
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = _font(); vc.border = _border(_C["mint"])

    # ── KPIs ──────────────────────────────────────────────────────────────────
    sec("D5", "📊  Key Performance Indicators")
    kpis = [
        ("Total Tests", total,        _C["blue"]),
        ("✅  Passed",   passed,       _C["green"]),
        ("❌  Failed",   failed,       _C["red"] if failed else _C["green"]),
        ("Pass Rate",   f"{rate}%",   _C["amber"] if float(rate) < 80 else _C["green"]),
    ]
    for i, (label, val, color) in enumerate(kpis):
        r = 6 + i
        ws.row_dimensions[r].height = 22
        lc = ws.cell(row=r, column=4, value=label)
        lc.font = _font(bold=True); lc.fill = _fill(_C["grey"]); lc.border = _border(_C["grey_bdr"])
        vc = ws.cell(row=r, column=5, value=val)
        vc.font = _font(size=13, bold=True, color=color)
        vc.alignment = Alignment(horizontal="center"); vc.border = _border(_C["grey_bdr"])

    # ── Category Status ──────────────────────────────────────────────────────
    sec("G5", "🏁  Category Status & Deployability")
    c_headers = ["Category", "Tests", "Pass", "Fail", "Status", "Deploy?"]
    for ci, h in enumerate(c_headers):
        c = ws.cell(row=6, column=7 + ci, value=h)
        c.font = _font(bold=True, color=_C["white"]); c.fill = _fill(_C["green"])
        c.border = _border(_C["mint"]); c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[6].height = 22

    g_total = g_pass = g_fail = 0
    for idx, (label, _key, cat_steps) in enumerate(cats):
        p = sum(1 for s in cat_steps if s.get("status") == "PASS")
        f = len(cat_steps) - p
        ok = f == 0 and len(cat_steps) > 0
        g_total += len(cat_steps); g_pass += p; g_fail += f
        r = 7 + idx
        vals = [label, len(cat_steps), p, f, "PASS" if ok else "FAIL", "DEPLOYABLE" if ok else "BLOCKED"]
        for ci, v in enumerate(vals):
            c = ws.cell(row=r, column=7 + ci, value=v)
            c.border = _border(_C["row_bdr"])
            c.alignment = Alignment(horizontal="left" if ci == 0 else "center", vertical="center")
            if ci >= 4:
                c.fill = _fill(_C["lt_green"] if ok else _C["lt_red"])
                c.font = _font(bold=True, color=_C["green"] if ok else _C["red"])
            else:
                c.fill = _fill(_C["white"] if idx % 2 == 0 else _C["grey"])
                c.font = _font()
        ws.row_dimensions[r].height = 19

    # Overall row
    tr = 7 + len(cats)
    ok_all = g_fail == 0
    for ci, v in enumerate(["OVERALL", g_total, g_pass, g_fail,
                             "PASS" if ok_all else "FAIL",
                             "DEPLOYABLE" if ok_all else "BLOCKED"]):
        c = ws.cell(row=tr, column=7 + ci, value=v)
        c.font = _font(bold=True, color=(_C["green"] if ok_all else _C["red"]) if ci >= 4 else _C["text"])
        c.fill = _fill(_C["grey"]); c.border = _border(_C["grey_bdr"])
        c.alignment = Alignment(horizontal="left" if ci == 0 else "center")
    ws.row_dimensions[tr].height = 22

    # Note
    nr = tr + 2
    ws.merge_cells(f"A{nr}:K{nr+1}")
    nc = ws[f"A{nr}"]
    nc.value = "📌  This report is auto-generated by the Green Harvest Buddy Appium E2E Suite. It contains 400 test cases across 4 categories. Refer to individual tabs for step-level details, screenshots, expected/actual values, and timestamps."
    nc.font = _font(size=9, italic=True, color=_C["sub"])
    nc.alignment = Alignment(wrap_text=True, vertical="top")

    for ci, w in enumerate([28, 36, 4, 22, 14, 4, 28, 10, 10, 10, 16, 16]):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w

    # ══════════════════════════════════════════════════════════════════════════
    # TABS 2-5 — DETAIL LOGS
    # ══════════════════════════════════════════════════════════════════════════
    headers = [
        ("Test ID",          14),
        ("Module / Feature", 22),
        ("Description",      42),
        ("Action Taken",     40),
        ("Expected Outcome", 40),
        ("Actual Result",    48),
        ("Status",           12),
        ("Duration (ms)",    14),
        ("Timestamp",        22),
    ]

    sheets = [
        ("UI-UX Tests",      groups["ui"]),
        ("Functional Tests", groups["func"]),
        ("Unit Tests",       groups["unit"]),
        ("Validation Tests", groups["val"]),
    ]

    for name, sheet_steps in sheets:
        wd = wb.create_sheet(name)

        # Banner
        wd.merge_cells("A1:I1")
        wd.row_dimensions[1].height = 28
        bn = wd["A1"]
        bn.value = f"📱  Green Harvest Buddy — {name}  ({len(sheet_steps)} cases)"
        bn.font = _font(size=12, bold=True, color=_C["white"])
        bn.fill = _fill(_C["green"])
        bn.alignment = Alignment(horizontal="left", vertical="center")

        # Header row
        wd.row_dimensions[2].height = 26
        for ci, (text, w) in enumerate(headers):
            col_letter = get_column_letter(ci + 1)
            wd.column_dimensions[col_letter].width = w
            c = wd.cell(row=2, column=ci + 1, value=text)
            c.font = _font(bold=True, color=_C["white"])
            c.fill = _fill(_C["dk_green"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = _med_border()

        # Data rows
        for ri, step in enumerate(sheet_steps):
            rn = ri + 3
            wd.row_dimensions[rn].height = 22
            alt = ri % 2 != 0
            bg = _C["grey"] if alt else _C["white"]

            vals = [
                step.get("id", ""),
                step.get("module", ""),
                step.get("description", ""),
                step.get("action", ""),
                step.get("expected", ""),
                step.get("actual", ""),
                step.get("status", ""),
                step.get("duration", 0),
                (step.get("timestamp", "") or "")[11:19],
            ]

            for ci, v in enumerate(vals):
                c = wd.cell(row=rn, column=ci + 1, value=v)
                c.font = _font()
                c.border = _border(_C["row_bdr"])
                c.alignment = Alignment(vertical="center", wrap_text=2 <= ci <= 5)
                if ci == 0:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 7:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = "#,##0"
                if ci == 8:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                if ci == 6:
                    ok = step.get("status") == "PASS"
                    c.fill = _fill(_C["lt_green"] if ok else _C["lt_red"])
                    c.font = _font(bold=True, color=_C["green"] if ok else _C["red"])
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.fill = _fill(bg)



    # Save
    wb.save(output_path)
    print(f"[+] Excel report → {output_path}")
